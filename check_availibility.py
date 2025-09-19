"""
Solar Plant Availability Calculator - Fixed Version
===================================================
Calculates daily availability for solar plants at different levels:
- Plant, Inverter, MPPT, and String levels
- Formula A: Time-based | Formula B: Irradiance-weighted
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import numpy as np

# ================================================================================================
# UTILITY FUNCTIONS
# ================================================================================================

def get_availability_color(value):
    """Return color code based on availability percentage"""
    if value == "Data Unavailable" or pd.isna(value):
        return "808080"  # Gray
    elif value == 100:
        return "00FF00"  # Green
    elif 98 <= value < 100:
        return "0000FF"  # Blue
    elif 95 <= value < 98:
        return "FFFF00"  # Yellow
    elif 80 <= value < 95:
        return "FFA500"  # Orange
    elif value < 80:
        return "FF0000"  # Red
    else:
        return "FFFFFF"

def apply_excel_coloring(excel_file, availability_column="Availability"):
    """Apply color coding to Excel availability column"""
    try:
        wb = load_workbook(excel_file)
        ws = wb.active
        df_temp = pd.read_excel(excel_file)
        col_idx = df_temp.columns.get_loc(availability_column) + 1

        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=col_idx)
            try:
                cell_value = float(cell.value)
            except:
                cell_value = cell.value
            color = get_availability_color(cell_value)
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

        wb.save(excel_file)
        print(f"✅ Saved: {excel_file}")
    except Exception as e:
        print(f"❌ Error saving {excel_file}: {e}")

def calculate_availability_percentage(numerator, denominator):
    """Calculate availability percentage"""
    if denominator == 0 or pd.isna(denominator):
        return "Data Unavailable"
    if pd.isna(numerator):
        numerator = 0
    return round((numerator / denominator * 100), 2)

# ================================================================================================
# PLANT LEVEL AVAILABILITY
# ================================================================================================

def calculate_plant_availability(df, plant_name, formula="A", irradiance_threshold=0.05,
                               excel_file="plant_availability.xlsx"):
    """Calculate daily availability for plant-level dataset (pre-filtered data)"""
    print(f"🏭 Processing Plant: {plant_name} | Formula: {formula}")

    if df.empty:
        print("❌ No data available")
        return pd.DataFrame(), pd.DataFrame()

    plant_df = df.copy()

    # Ensure timestamp is datetime format
    plant_df["timestamp"] = pd.to_datetime(plant_df["timestamp"])

    # Calculate base conditions
    plant_df["Condition_Numerator"] = (
        (plant_df["dataItemMap.radiation_intensity"] > irradiance_threshold) &
        (plant_df["dataItemMap.inverter_power"] > 0)
    ).astype(int)

    plant_df["Condition_Denominator"] = (
        plant_df["dataItemMap.radiation_intensity"] > irradiance_threshold
    ).astype(int)

    plant_df["Potential_Weight"] = plant_df["dataItemMap.radiation_intensity"].where(
        plant_df["dataItemMap.radiation_intensity"] > irradiance_threshold, 0
    )

    plant_df["Actual_Weight"] = plant_df["dataItemMap.radiation_intensity"].where(
        (plant_df["dataItemMap.radiation_intensity"] > irradiance_threshold) &
        (plant_df["dataItemMap.inverter_power"] > 0), 0
    )

    # Aggregate by date
    plant_df["Date"] = plant_df["timestamp"].dt.date
    daily = plant_df.groupby("Date").agg({
        "Condition_Numerator": "sum",
        "Condition_Denominator": "sum",
        "Actual_Weight": "sum",
        "Potential_Weight": "sum"
    }).reset_index()

    # Calculate availability and select appropriate columns based on formula
    if formula == "A":
        daily["Availability"] = daily.apply(
            lambda row: calculate_availability_percentage(row["Condition_Numerator"], row["Condition_Denominator"]),
            axis=1
        )
        daily.insert(1, "Plant", plant_name)
        daily_availability = daily[["Date", "Plant", "Condition_Numerator", "Condition_Denominator", "Availability"]]
    elif formula == "B":
        daily["Availability"] = daily.apply(
            lambda row: calculate_availability_percentage(row["Actual_Weight"], row["Potential_Weight"]),
            axis=1
        )
        daily.insert(1, "Plant", plant_name)
        daily_availability = daily[["Date", "Plant", "Actual_Weight", "Potential_Weight", "Availability"]]

    # Create debug dataframe
    debug_df = plant_df[["Plant", "timestamp", "dataItemMap.inverter_power", "dataItemMap.radiation_intensity",
                         "Condition_Numerator", "Condition_Denominator", "Potential_Weight", "Actual_Weight"]]

    # Save to Excel with coloring
    daily_availability.to_excel(excel_file, index=False)
    apply_excel_coloring(excel_file)

    return daily_availability, debug_df

# ================================================================================================
# INVERTER LEVEL AVAILABILITY
# ================================================================================================

def calculate_inverter_availability(df, plant_name, formula="A", irradiance_threshold=0.05,
                                  excel_file="inverter_availability.xlsx"):
    """Calculate daily availability for inverter-level dataset (pre-filtered data)"""
    print(f"🔌 Processing Inverters: {plant_name} | Formula: {formula}")

    if df.empty:
        print("❌ No data available")
        return pd.DataFrame(), pd.DataFrame()

    plant_df = df.copy()

    # Ensure timestamp is datetime format
    plant_df["timestamp"] = pd.to_datetime(plant_df["timestamp"])

    # Create unique inverter ID
    plant_df["sn_id"] = plant_df["Plant"].astype(str) + "_" + plant_df["sn"].astype(str)

    # Calculate base conditions
    plant_df["Condition_Numerator"] = (
        (plant_df["radiation_intensity"] > irradiance_threshold) &
        (plant_df["InverterPower"] > 0)
    ).astype(int)

    plant_df["Condition_Denominator"] = (
        plant_df["radiation_intensity"] > irradiance_threshold
    ).astype(int)

    plant_df["Potential_Weight"] = plant_df["radiation_intensity"].where(
        plant_df["radiation_intensity"] > irradiance_threshold, 0
    )

    plant_df["Actual_Weight"] = plant_df["radiation_intensity"].where(
        (plant_df["radiation_intensity"] > irradiance_threshold) &
        (plant_df["InverterPower"] > 0), 0
    )

    # Aggregate by date and inverter
    plant_df["Date"] = plant_df["timestamp"].dt.date
    daily = plant_df.groupby(["Date", "sn_id"]).agg({
        "Condition_Numerator": "sum",
        "Condition_Denominator": "sum",
        "Actual_Weight": "sum",
        "Potential_Weight": "sum"
    }).reset_index()

    # Calculate availability and select appropriate columns based on formula
    if formula == "A":
        daily["Availability"] = daily.apply(
            lambda row: calculate_availability_percentage(row["Condition_Numerator"], row["Condition_Denominator"]),
            axis=1
        )
        daily.insert(1, "Plant", plant_name)
        daily_availability = daily[["Date", "Plant", "sn_id", "Condition_Numerator", "Condition_Denominator", "Availability"]]
    elif formula == "B":
        daily["Availability"] = daily.apply(
            lambda row: calculate_availability_percentage(row["Actual_Weight"], row["Potential_Weight"]),
            axis=1
        )
        daily.insert(1, "Plant", plant_name)
        daily_availability = daily[["Date", "Plant", "sn_id", "Actual_Weight", "Potential_Weight", "Availability"]]

    # Create debug dataframe
    debug_df = plant_df[["Plant", "sn_id", "timestamp", "InverterPower", "radiation_intensity",
                         "Condition_Numerator", "Condition_Denominator", "Potential_Weight", "Actual_Weight"]]

    # Save to Excel with coloring
    daily_availability.to_excel(excel_file, index=False)
    apply_excel_coloring(excel_file)

    return daily_availability, debug_df

# ================================================================================================
# MPPT LEVEL AVAILABILITY
# ================================================================================================

def calculate_mppt_availability(df, plant_name, formula="A", irradiance_threshold=0.05,
                              excel_file=None):
    """Calculate daily availability at MPPT level (pre-filtered data)"""
    print(f"⚡ Processing MPPTs: {plant_name} | Formula: {formula}")

    if df.empty:
        print("❌ No data available")
        return pd.DataFrame(), pd.DataFrame()

    plant_df = df.copy()

    # Ensure timestamp is datetime format
    plant_df["timestamp"] = pd.to_datetime(plant_df["timestamp"])

    # Identify MPPT columns
    mppt_power_cols = [col for col in df.columns if 'mppt_' in col.lower() and 'capacity' not in col.lower()]
    mppt_id_cols = [col for col in df.columns if 'mppt' in col.lower() and 'mpptid' in col.lower()]

    # Transform data to long format
    records = []
    for _, row in plant_df.iterrows():
        sn = row['sn']
        timestamp = row['timestamp']
        radiation = row['radiation_intensity']

        for i, mppt_col in enumerate(mppt_power_cols):
            if i < len(mppt_id_cols):
                mppt_power = row[mppt_col]
                mppt_name = row[mppt_id_cols[i]]
                mppt_id = f"{plant_name}_{sn}_{mppt_name}".replace(" ", "_")

                records.append({
                    "timestamp": timestamp,
                    "plant": plant_name,
                    "sn": sn,
                    "mppt": mppt_name,
                    "mppt_id": mppt_id,
                    "InverterPower": mppt_power,
                    "radiation_intensity": radiation
                })

    mppt_df = pd.DataFrame(records)

    # Calculate base conditions
    mppt_df["Condition_Numerator"] = (
        (mppt_df["radiation_intensity"] > irradiance_threshold) &
        (mppt_df["InverterPower"] > 0)
    ).astype(int)

    mppt_df["Condition_Denominator"] = (
        mppt_df["radiation_intensity"] > irradiance_threshold
    ).astype(int)

    mppt_df["Potential_Weight"] = mppt_df["radiation_intensity"].where(
        mppt_df["radiation_intensity"] > irradiance_threshold, 0
    )

    mppt_df["Actual_Weight"] = mppt_df["radiation_intensity"].where(
        (mppt_df["radiation_intensity"] > irradiance_threshold) &
        (mppt_df["InverterPower"] > 0), 0
    )

    # Aggregate by date and MPPT
    mppt_df["Date"] = mppt_df["timestamp"].dt.date
    daily = mppt_df.groupby(["Date", "mppt_id"]).agg({
        "Condition_Numerator": "sum",
        "Condition_Denominator": "sum",
        "Actual_Weight": "sum",
        "Potential_Weight": "sum"
    }).reset_index()

    # Calculate availability and select appropriate columns based on formula
    if formula == "A":
        daily["Availability"] = daily.apply(
            lambda row: calculate_availability_percentage(row["Condition_Numerator"], row["Condition_Denominator"]),
            axis=1
        )
        output_columns = ["Date", "mppt_id", "Condition_Numerator", "Condition_Denominator", "Availability"]
    elif formula == "B":
        daily["Availability"] = daily.apply(
            lambda row: calculate_availability_percentage(row["Actual_Weight"], row["Potential_Weight"]),
            axis=1
        )
        output_columns = ["Date", "mppt_id", "Actual_Weight", "Potential_Weight", "Availability"]

    daily_availability = daily[output_columns].copy()

    # Save to Excel with coloring
    if excel_file is None:
        excel_file = f"{plant_name.replace(' ', '_')}_mppt_daily_{formula}.xlsx"

    daily_availability.to_excel(excel_file, index=False)
    apply_excel_coloring(excel_file)

    return daily_availability, mppt_df

# ================================================================================================
# STRING LEVEL AVAILABILITY
# ================================================================================================

def calculate_string_availability(df, plant_name, formula="A", irradiance_threshold=0.05,
                                     excel_file=None):
    """Calculate daily availability at string level from HR_IL_PRD data"""
    print(f"🔗 Processing HR_IL_PRD Strings: {plant_name} | Formula: {formula}")

    if df.empty:
        print("❌ No data available")
        return pd.DataFrame(), pd.DataFrame()

    plant_df = df.copy()

    # --- Ensure timestamp from Day_Hour ---
    if "timestamp" not in plant_df.columns:
        plant_df["timestamp"] = pd.to_datetime(plant_df["Day_Hour"], errors="coerce")
    else:
        plant_df["timestamp"] = pd.to_datetime(plant_df["timestamp"], errors="coerce")

    # ✅ Keep only configured strings
    plant_df = plant_df[plant_df["String_Configured"] == 1].copy()

    if plant_df.empty:
        print("⚠️ No configured strings found (String_Configured == 1)")
        return pd.DataFrame(), df

    # --- Unique string ID ---
    plant_df["string_id"] = (
        plant_df["Plant"].astype(str) + "_" +
        plant_df["sn"].astype(str) + "_" +
        plant_df["MPPT"].astype(str) + "_" +
        plant_df["Strings"].astype(str)
    )

    # --- Base conditions ---
    plant_df["Condition_Numerator"] = (
        (plant_df["radiation_intensity"] > irradiance_threshold) &
        (plant_df["Watt/String"] > 0)
    ).astype(int)

    plant_df["Condition_Denominator"] = (
        plant_df["radiation_intensity"] > irradiance_threshold
    ).astype(int)

    plant_df["Actual_Weight"] = plant_df["radiation_intensity"].where(
        (plant_df["radiation_intensity"] > irradiance_threshold) &
        (plant_df["Watt/String"] > 0), 0
    )

    plant_df["Potential_Weight"] = plant_df["radiation_intensity"].where(
        plant_df["radiation_intensity"] > irradiance_threshold, 0
    )

    # --- Aggregate by Date + String ---
    plant_df["Date"] = plant_df["timestamp"].dt.date

    if formula.upper() == "A":
        daily = plant_df.groupby(["Date", "string_id"]).agg({
            "Condition_Numerator": "sum",
            "Condition_Denominator": "sum"
        }).reset_index()

        daily["Availability"] = daily.apply(
            lambda row: calculate_availability_percentage(row["Condition_Numerator"],
                                                          row["Condition_Denominator"]),
            axis=1
        )
        output_columns = ["Date", "string_id", "Condition_Numerator", "Condition_Denominator", "Availability"]

    elif formula.upper() == "B":
        daily = plant_df.groupby(["Date", "string_id"]).agg({
            "Actual_Weight": "sum",
            "Potential_Weight": "sum"
        }).reset_index()

        daily["Availability"] = daily.apply(
            lambda row: calculate_availability_percentage(row["Actual_Weight"],
                                                          row["Potential_Weight"]),
            axis=1
        )
        output_columns = ["Date", "string_id", "Actual_Weight", "Potential_Weight", "Availability"]

    else:
        raise ValueError("❌ Invalid formula. Use 'A' or 'B'.")

    # --- Safe selection ---
    daily_availability = daily[output_columns].copy()

    # --- Save Excel ---
    if excel_file is None:
        excel_file = f"{plant_name.replace(' ', '_')}_HR_string_daily_{formula}.xlsx"

    daily_availability.to_excel(excel_file, index=False)
    apply_excel_coloring(excel_file)

    return daily_availability, plant_df



# ================================================================================================
# MAIN CALCULATION FUNCTION
# ================================================================================================

def calculate_availability_main(df, plant_name, level, formula="A", irradiance_threshold=0.05, excel_file=None):
    """
    Main function to calculate availability at different levels

    Args:
        df: Pre-filtered pandas DataFrame (plant and date range already filtered)
        plant_name: Plant name (for naming/identification only)
        level: 'plant', 'inverter', 'mppt', 'string'
        formula: 'A' (time-based) or 'B' (irradiance-weighted)
        irradiance_threshold: Minimum irradiance threshold (default: 0.05)
        excel_file: Output Excel file path

    Returns:
        tuple: (daily_availability_df, debug_df)
    """
    print(f"🚀 Starting {level.upper()} level calculation | Plant: {plant_name} | Formula: {formula}")

    level = level.lower()
    formula = formula.upper()

    if level not in ["plant", "inverter", "mppt", "string"]:
        raise ValueError("Level must be: 'plant', 'inverter', 'mppt', 'string'")
    if formula not in ["A", "B"]:
        raise ValueError("Formula must be 'A' or 'B'")

    # Generate default Excel filename if not provided
    if excel_file is None:
        excel_file = f"{plant_name.replace(' ', '_')}_{level}_daily_{formula}.xlsx"

    # Call appropriate calculation function based on level
    if level == "plant":
        daily_availability, debug_df = calculate_plant_availability(
            df=df, plant_name=plant_name, formula=formula,
            irradiance_threshold=irradiance_threshold, excel_file=excel_file
        )
    elif level == "inverter":
        daily_availability, debug_df = calculate_inverter_availability(
            df=df, plant_name=plant_name, formula=formula,
            irradiance_threshold=irradiance_threshold, excel_file=excel_file
        )
    elif level == "mppt":
        daily_availability, debug_df = calculate_mppt_availability(
            df=df, plant_name=plant_name, formula=formula,
            irradiance_threshold=irradiance_threshold, excel_file=excel_file
        )
    elif level == "string":
        daily_availability, debug_df = calculate_string_availability(
            df=df, plant_name=plant_name, formula=formula,
            irradiance_threshold=irradiance_threshold, excel_file=excel_file
        )

    # Final summary
    if not daily_availability.empty:
        total_records = len(daily_availability)
        numeric_availability = pd.to_numeric(daily_availability['Availability'], errors='coerce')
        valid_data = numeric_availability.dropna()

        if len(valid_data) > 0:
            print(f"📈 Results: {total_records} records | Avg: {valid_data.mean():.1f}% | Range: {valid_data.min():.1f}%-{valid_data.max():.1f}%")

    return daily_availability, debug_df
# #
# if __name__ == "__main__":
#     df = pd.read_csv('/content/string_level_with_radiation.csv')
#     results, debug = calculate_availability_main(df, plant_name='Coca Cola Faisalabad',level='string',formula='A')