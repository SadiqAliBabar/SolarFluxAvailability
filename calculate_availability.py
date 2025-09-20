import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import argparse
import logging
from fetch_data import fetch_plant_data, fetch_inverter_data, fetch_mppt_data, fetch_string_data, get_plant_names

logging.basicConfig(level=logging.INFO)

def get_availability_color(value):
    """Color based on availability %."""
    if pd.isna(value) or value == "Data Unavailable":
        return "808080"
    value = float(value)
    if value == 100:
        return "00FF00"
    elif 98 <= value < 100:
        return "0000FF"
    elif 95 <= value < 98:
        return "FFFF00"
    elif 80 <= value < 95:
        return "FFA500"
    elif value < 80:
        return "FF0000"
    return "FFFFFF"

def apply_coloring(excel_file, col_name='Availability'):
    """Apply colors to Excel column."""
    try:
        wb = load_workbook(excel_file)
        ws = wb.active
        df = pd.read_excel(excel_file)
        col_idx = df.columns.get_loc(col_name) + 1
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=col_idx)
            color = get_availability_color(cell.value)
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        wb.save(excel_file)
        logging.info(f"Saved colored Excel: {excel_file}")
    except Exception as e:
        logging.error(f"Error coloring {excel_file}: {e}")

def calculate_availability(df, level, formula='A', irradiance_threshold=0.05, power_threshold=0.0):
    """Core availability calc for any level with customizable power threshold."""
    if df.empty:
        return pd.DataFrame(), pd.DataFrame()

    df = df.copy()
    # Select power and radiation columns based on level
    power_col = ('dataItemMap.inverter_power' if level == 'plant' else
                 'InverterPower' if level == 'inverter' else
                 'mppt_Power' if level == 'mppt' else
                 'P_abd')  # Use P_abd for string level
    rad_col = 'dataItemMap.radiation_intensity' if level == 'plant' else 'radiation_intensity'
    id_col = ('sn' if level == 'inverter' else
              ['sn', 'mpptId'] if level == 'mppt' else
              ['sn', 'MPPT', 'Strings'] if level == 'string' else
              None)

    df['timestamp'] = pd.to_datetime(df.get('timestamp', df.get('Day_Hour')), errors='coerce')
    df['Date'] = df['timestamp'].dt.date

    if level == 'string':
        df = df[df['String_Configured'] == 1]

    # Ensure required columns exist
    if power_col not in df.columns or rad_col not in df.columns:
        logging.error(f"Missing required columns: {power_col} or {rad_col}")
        return pd.DataFrame(), pd.DataFrame()

    # Use power_threshold instead of hardcoded power > 0
    df['Num'] = ((df[rad_col] > irradiance_threshold) & (df[power_col] > power_threshold)).astype(int)
    df['Den'] = (df[rad_col] > irradiance_threshold).astype(int)
    df['Act_Wt'] = np.where((df[rad_col] > irradiance_threshold) & (df[power_col] > power_threshold), df[rad_col], 0)
    df['Pot_Wt'] = np.where(df[rad_col] > irradiance_threshold, df[rad_col], 0)

    group_cols = ['Date']
    if id_col:
        if isinstance(id_col, list):
            group_cols.extend(id_col)
        else:
            group_cols.append(id_col)
    if 'Plant' in df.columns:
        group_cols.append('Plant')

    daily = df.groupby(group_cols).agg({
        'Num': 'sum', 'Den': 'sum', 'Act_Wt': 'sum', 'Pot_Wt': 'sum'
    }).reset_index()

    if formula == 'A':
        daily['Availability'] = np.round((daily['Num'] / daily['Den'].replace(0, np.nan)) * 100, 2).fillna('Data Unavailable')
        if level == 'mppt':
            daily = daily[['Date', 'Plant', 'sn', 'mpptId', 'Num', 'Den', 'Availability']]
        elif level == 'string':
            daily = daily[['Date', 'Plant', 'sn', 'MPPT', 'Strings', 'Num', 'Den', 'Availability']]
        else:
            daily = daily[group_cols + ['Num', 'Den', 'Availability']]
    else:  # B
        daily['Availability'] = np.round((daily['Act_Wt'] / daily['Pot_Wt'].replace(0, np.nan)) * 100, 2).fillna('Data Unavailable')
        if level == 'mppt':
            daily = daily[['Date', 'Plant', 'sn', 'mpptId', 'Act_Wt', 'Pot_Wt', 'Availability']]
        elif level == 'string':
            daily = daily[['Date', 'Plant', 'sn', 'MPPT', 'Strings', 'Act_Wt', 'Pot_Wt', 'Availability']]
        else:
            daily = daily[group_cols + ['Act_Wt', 'Pot_Wt', 'Availability']]

    debug_df = df[['Plant', 'timestamp', power_col, rad_col, 'Num', 'Den', 'Act_Wt', 'Pot_Wt']] if 'Plant' in df.columns else df[['timestamp', power_col, rad_col, 'Num', 'Den', 'Act_Wt', 'Pot_Wt']]
    return daily, debug_df

def main():
    parser = argparse.ArgumentParser(description="Calculate solar availability.")
    parser.add_argument('--level', required=True, choices=['plant', 'inverter', 'mppt', 'string'])
    parser.add_argument('--start_date')
    parser.add_argument('--end_date')
    parser.add_argument('--plant_name', default='all', help="Plant name, 'all', or comma-separated list (e.g., 'Coca_Cola_Faisalabad,Metro_DHA')")
    parser.add_argument('--inverter_sn', default='all', help="Inverter serial number, 'all', or comma-separated list (e.g., '6T21B9040017,6T21B9040018')")
    parser.add_argument('--mppt_id', default='all', help="MPPT ID, 'all', or comma-separated list (e.g., 'MPPT1,MPPT5')")
    parser.add_argument('--string_id', default='all', help="String ID, 'all', or comma-separated list (e.g., 'pv9,pv23')")
    parser.add_argument('--connection_string', default='mongodb://110.39.23.106:27023/')
    parser.add_argument('--formula', default='A', choices=['A', 'B'])
    parser.add_argument('--irradiance_threshold', type=float, default=0.05)
    parser.add_argument('--power_threshold', type=float, default=0.0, help="Power threshold for availability calculation (default: 0.0)")
    parser.add_argument('--output_excel', default=None, help="Output Excel file (default: {plant}_{level}_{formula}_availability.xlsx)")

    args = parser.parse_args()

    # Set default output_excel based on plant_name, level, and formula
    if args.output_excel is None:
        if args.plant_name.lower() == 'all':
            plant_part = 'all_plants'
        elif ',' in args.plant_name:
            plant_part = 'multiple_plants'
        else:
            plant_part = args.plant_name.replace(' ', '_')
        args.output_excel = f"{plant_part}_{args.level}_{args.formula}_availability.xlsx"

    logging.info(f"Starting calc for {args.level} | Plant(s): {args.plant_name} | Inverter(s): {args.inverter_sn} | MPPT(s): {args.mppt_id} | String(s): {args.string_id} | Formula: {args.formula} | Irradiance Threshold: {args.irradiance_threshold} | Power Threshold: {args.power_threshold} | Output: {args.output_excel}")

    if args.plant_name.lower() == 'all':
        plants = get_plant_names(args.connection_string)
    else:
        plants = [p.strip().replace(' ', '_') for p in args.plant_name.split(',')]

    if not plants:
        logging.error("No plants found.")
        return

    all_dfs = []
    for plant in plants:
        plant_display = plant.replace('_', ' ')
        logging.info(f"Processing plant: {plant_display}")

        if args.level == 'plant':
            df = fetch_plant_data(args.connection_string, plant, args.start_date, args.end_date)
        elif args.level == 'inverter':
            df = fetch_inverter_data(args.connection_string, plant, args.inverter_sn, args.start_date, args.end_date)
        elif args.level == 'mppt':
            df = fetch_mppt_data(args.connection_string, plant, args.inverter_sn, args.mppt_id, args.start_date, args.end_date)
        elif args.level == 'string':
            df = fetch_string_data(args.connection_string, plant, args.inverter_sn, args.mppt_id, args.string_id, args.start_date, args.end_date)
        else:
            df = pd.DataFrame()

        if df.empty:
            logging.warning(f"No data for {plant_display} at {args.level} level.")
            continue

        # Clean numerics
        numeric_cols = ['dataItemMap.inverter_power', 'dataItemMap.radiation_intensity', 'InverterPower', 'radiation_intensity', 'mppt_Power', 'P_abd']
        for col in numeric_cols:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors='coerce')

        daily, debug = calculate_availability(df, args.level, args.formula, args.irradiance_threshold, args.power_threshold)
        if 'Plant' not in daily.columns:
            daily.insert(0, 'Plant', plant_display)
        all_dfs.append(daily)

        # Optional debug CSV
        # debug.to_csv(f"debug_{plant}_{args.level}.csv", index=False)

    if all_dfs:
        final_df = pd.concat(all_dfs, ignore_index=True)
        final_df.to_excel(args.output_excel, index=False)
        apply_coloring(args.output_excel)
        logging.info(f"Results saved to {args.output_excel}")
        print(final_df.head())
    else:
        logging.error("No results generated.")

if __name__ == '__main__':
    main()
